"""
Excel 渲染（design.md §7.4 步骤 4-5）。纯数据对比表，无图表、无结论。

版式（矩阵式宽表）：
  - 行 = (Input_Length, Concurrency) 组合，按输入长度分块
  - 列 = 每个指标一个"列组"，组内 = 对比轴各取值 + 两两差异列（≥2 个取值时）
  - 双层表头：第 1 行为指标名（跨列合并），第 2 行为对比轴取值 / "A vs B"
  - 每个 Input_Length 块结束后插入一行块汇总：该块内各差异列的均值
"""
from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from autores.server.report import hardware

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=12)
_DIM_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_DELTA_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_DELTA_POS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_DELTA_NEG_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 差异列的数字格式：带符号的百分比
_DELTA_FORMAT = "0.00%"


def _constraints_text(constraints: dict) -> str:
    # 维度名 -> 中文展示
    label_map = {
        "model": "模型", "model_version": "模型版本",
        "framework": "框架", "framework_version": "框架版本",
        "gpu_type": "显卡", "tp": "TP", "dp": "DP", "pp": "PP",
        "ep": "EP", "cp": "CP", "kv_cache_dtype": "KV Cache Dtype",
        "hicache_enabled": "HiCache", "flexkv_enabled": "FlexKV",
        "torch_compile": "TorchCompile", "quantization": "量化",
        "attention_backend": "Attention后端",
    }
    parts = [f"{label_map.get(k, k)}: {v}" for k, v in constraints.items()]
    return " | ".join(parts)


def _as_number(value):
    """单元格值转 float；N/A、None、非数值返回 None。"""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


# 延迟类指标：越低越好，差异列符号与吞吐类相反
_LOWER_IS_BETTER_PREFIXES = ("TTFT_", "TPOT_", "ITL_", "E2E_")


def _is_lower_better(metric_name: str) -> bool:
    return metric_name.startswith(_LOWER_IS_BETTER_PREFIXES)


def _relative_delta(base, ref, *, lower_is_better: bool = False):
    """
    相对差异 (A vs B)，A=base（对比轴第一列），B=ref（第二列）。
    吞吐类：越高越好 → (A-B)/B，A 更高则为正。
    延迟类：越低越好 → (B-A)/B，A 更低则为正。
    """
    b, r = _as_number(base), _as_number(ref)
    if b is None or r is None or r == 0:
        return None
    if lower_is_better:
        return (r - b) / r
    return (b - r) / r


def _compare_pairs(column_labels: list) -> list[tuple]:
    """对比轴取值的两两组合 (A, B, "A vs B")，A 在前、B 在后。"""
    pairs = []
    for i, a in enumerate(column_labels):
        for b in column_labels[i + 1:]:
            pairs.append((a, b, f"{a} vs {b}"))
    return pairs


def _write_delta_cell(ws, row: int, col: int, delta, accum: dict[int, list[float]]):
    """写入单个差异单元格并累计块汇总。"""
    c = ws.cell(row=row, column=col, value=delta if delta is not None else "N/A")
    c.alignment = _CENTER
    c.border = _BORDER
    if delta is not None:
        c.number_format = _DELTA_FORMAT
        _style_delta_cell(c, delta)
        accum[col].append(delta)
    else:
        c.fill = _DELTA_FILL


def _style_delta_cell(cell, delta, *, bold: bool = False):
    """正数绿色、负数红色；零或 N/A 不着色。"""
    if not isinstance(delta, (int, float)):
        return
    if delta > 0:
        cell.font = Font(bold=bold, color="006100")
        cell.fill = _DELTA_POS_FILL
    elif delta < 0:
        cell.font = Font(bold=bold, color="C00000")
        cell.fill = _DELTA_NEG_FILL


def render_comparison(table: dict, compare_on: str, output_dir: str) -> str:
    """把对比宽表渲染为 xlsx，返回文件路径。"""
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "对比报告"

    column_labels = list(table["column_labels"])
    matrix = table["matrix"]
    metric_names = table["metric_names"]

    gpu_scaled = table.get("gpu_scaled", False)
    column_gpus = table.get("column_gpus", {})
    column_scale = table.get("column_scale", {})

    def _col_display(label) -> str:
        """列标签：弱扩展换算时追加 (×比例) 标识。"""
        scale = column_scale.get(label, 1)
        if gpu_scaled and scale not in (None, 1):
            return f"{label} (×{scale:g})"
        return str(label)

    def _metric_display(mname: str) -> str:
        """吞吐类指标名：换算后追加 (× scale) 标识。"""
        if gpu_scaled and mname in hardware.THROUGHPUT_METRICS:
            return f"{mname} (× scale)"
        return mname

    compare_pairs = _compare_pairs(column_labels)
    # 每个指标占的列数：对比轴取值数 + 两两差异列数
    group_width = len(column_labels) + len(compare_pairs)

    # ── 标题区 ──
    ws.cell(row=1, column=1, value=f"对比轴: {compare_on}").font = _TITLE_FONT
    constraints = _constraints_text(table["constraints"])
    if constraints:
        ws.cell(row=2, column=1, value=f"约束: {constraints}")
    note_bits = []
    if table["notes"].get("multi_framework"):
        note_bits.append("含多个框架（版本号不可跨框架比较）")
    if table["notes"].get("multi_version"):
        note_bits.append("含多个框架版本（各自独立取出）")
    if len(column_labels) > 2:
        note_bits.append(f"对比轴 {len(column_labels)} 个取值，差异列为两两对比")
    if gpu_scaled:
        parts = []
        for label in column_labels:
            gpus = column_gpus.get(label)
            if gpus:
                scale = column_scale.get(label, 1)
                tag = f"×{scale:g}" if scale not in (None, 1) else "基准"
                parts.append(f"{label}={hardware.unit_desc(gpus)}({tag})")
        note_bits.append(
            "已按卡数弱扩展归一：吞吐类×卡数比例、concurrency 同比对齐、"
            "延迟类(TTFT/TPOT/ITL/E2E)保持原值　" + "，".join(parts)
        )
    if note_bits:
        ws.cell(row=3, column=1, value="说明: " + "；".join(note_bits))

    top_row, sub_row = 5, 6
    data_start = 7

    # ── 双层表头 ──
    # 左侧两列维度列：Input_Length / Concurrency，纵向合并两行
    for idx, dim_label in enumerate(("Input_Length", "Concurrency"), start=1):
        ws.merge_cells(start_row=top_row, start_column=idx, end_row=sub_row, end_column=idx)
        ws.cell(row=top_row, column=idx, value=dim_label)

    # 指标列组
    col = 3
    delta_columns: list[int] = []
    for mname in metric_names:
        end_col = col + group_width - 1
        if group_width > 1:
            ws.merge_cells(start_row=top_row, start_column=col, end_row=top_row, end_column=end_col)
        ws.cell(row=top_row, column=col, value=_metric_display(mname))
        for offset, col_label in enumerate(column_labels):
            ws.cell(row=sub_row, column=col + offset, value=_col_display(col_label))
        for k, (label_a, label_b, _) in enumerate(compare_pairs):
            delta_col = col + len(column_labels) + k
            ws.cell(row=sub_row, column=delta_col,
                    value=f"{_col_display(label_a)} vs {_col_display(label_b)}")
            delta_columns.append(delta_col)
        col = end_col + 1

    last_col = col - 1
    for r in (top_row, sub_row):
        for j in range(1, last_col + 1):
            c = ws.cell(row=r, column=j)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = _CENTER
            c.border = _BORDER

    # ── 数据区：按 Input_Length 分块，块尾追加差异均值汇总行 ──
    # matrix 已按 (input_length, concurrency) 排好序
    row_cursor = data_start
    delta_accum: dict[int, list[float]] = {j: [] for j in delta_columns}
    prev_input_length = None

    def flush_block(cursor: int) -> int:
        """写入当前块的差异均值汇总行，返回下一个可用行号。"""
        if not compare_pairs or not any(delta_accum.values()):
            delta_accum.update({j: [] for j in delta_columns})
            return cursor
        for j in delta_columns:
            vals = delta_accum[j]
            if not vals:
                continue
            avg = sum(vals) / len(vals)
            c = ws.cell(row=cursor, column=j, value=avg)
            c.alignment = _CENTER
            c.number_format = _DELTA_FORMAT
            c.border = _BORDER
            _style_delta_cell(c, avg, bold=True)
        delta_accum.update({j: [] for j in delta_columns})
        return cursor + 1

    for entry in matrix:
        il = entry["input_length"]
        if prev_input_length is not None and il != prev_input_length:
            row_cursor = flush_block(row_cursor)
        prev_input_length = il

        ws.cell(row=row_cursor, column=1, value=il)
        ws.cell(row=row_cursor, column=2, value=entry["concurrency"])
        for j in (1, 2):
            c = ws.cell(row=row_cursor, column=j)
            c.fill = _DIM_FILL
            c.alignment = _CENTER
            c.border = _BORDER

        col = 3
        for mname in metric_names:
            per_column = entry["metrics"].get(mname, {})
            for offset, col_label in enumerate(column_labels):
                c = ws.cell(row=row_cursor, column=col + offset,
                            value=per_column.get(col_label, "N/A"))
                c.alignment = _CENTER
                c.border = _BORDER
            lower = _is_lower_better(mname)
            for k, (label_a, label_b, _) in enumerate(compare_pairs):
                delta_col = col + len(column_labels) + k
                delta = _relative_delta(
                    per_column.get(label_a),
                    per_column.get(label_b),
                    lower_is_better=lower,
                )
                _write_delta_cell(ws, row_cursor, delta_col, delta, delta_accum)
            col += group_width
        row_cursor += 1

    if prev_input_length is not None:
        row_cursor = flush_block(row_cursor)

    # 冻结表头 + 左侧两个维度列
    ws.freeze_panes = ws.cell(row=data_start, column=3)

    # 列宽自适应（简单估算）
    for j in range(1, last_col + 1):
        letter = get_column_letter(j)
        max_len = 0
        for i in range(sub_row, row_cursor):
            v = ws.cell(row=i, column=j).value
            if v is not None:
                max_len = max(max_len, len(f"{v:.4f}" if isinstance(v, float) else str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 32)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"对比报告_{compare_on}_{ts}.xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)
    return path
