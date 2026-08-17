"""
手工上传入库（result.csv + 启动命令文本 + 表单元信息）。

场景：数据分散在不同子系统/地区，未落到 Scanner 扫描的 NAS 目录下，
由测试人员在前端页面直接提交一份整理好的 CSV 与启动命令文本。

与目录流的一致性保证：
  - CSV 解析复用 scanner.parser 的行→metric 逻辑（同样的列名归一与数值转换）；
  - 启动参数提取复用 tools/to_csv.py 的规则（见 launch_params 模块）；
  - 产出的文档结构与 parse_run_dir 完全一致，走同一个 db.insert_run；
  - 同时落盘到 config.scanner.benchmark_root，目录结构与 Scanner 一致，便于崩溃后重扫入库。
差异仅在于元信息来源：目录流读 metadata.json，上传流读表单字段。
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone

import importlib.util
import os
import threading

from autores.common.logging import get_logger
from autores.db import schema
from autores.scanner.parser import _to_number
from autores.server.ingest import launch_params
from autores.server.ingest.csv_columns import (
    SPEC_COLUMNS,
    build_header_map,
    check_required_dimensions,
    format_mapping_summary,
)
from autores.server.ingest import persist

log = get_logger("upload")

# 表单必填的元信息字段（metadata.json 在上传流里的替代品）
# 模型无版本时只填 model；model_version 入库写空串，不要求用户填写。
REQUIRED_META = ["framework", "framework_version", "model", "gpu_type"]

# 压测工具框架（bench framework）可选值。与 server framework 相互独立、禁止默认一致。
SUPPORTED_BENCH_FRAMEWORKS = ["sglang", "vllm"]

# CSV 单元格里视为"无值"的取值（判断 spec 列是否有值时用）；注意 "0" 算有值。
_NA_CELLS = frozenset({"", "n/a", "na", "none", "null"})


def supported_bench_frameworks() -> list[str]:
    return list(SUPPORTED_BENCH_FRAMEWORKS)


def _parse_bool_form(val) -> bool | None:
    """表单布尔值解析：识别 true/false/1/0/yes/no/on/off；无法识别返回 None。"""
    if isinstance(val, bool):
        return val
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("true", "1", "yes", "on"):
        return True
    if s in ("false", "0", "no", "off"):
        return False
    return None


def _cell_has_value(raw) -> bool:
    """CSV 单元格是否为真实取值（非空、非 N/A）；数值 0 视为有值。"""
    if raw is None:
        return False
    s = str(raw).strip()
    return bool(s) and s.lower() not in _NA_CELLS

_GPU_PRESETS = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
    "tools", "gpu_memory_presets.py",
)
_gpu_module = None
_gpu_lock = threading.Lock()


def _load_gpu_presets():
    """按路径加载 tools/gpu_memory_presets.py，显卡下拉与校验共用同一份型号表。"""
    global _gpu_module
    if _gpu_module is not None:
        return _gpu_module
    with _gpu_lock:
        if _gpu_module is not None:
            return _gpu_module
        spec = importlib.util.spec_from_file_location("_autores_gpu_presets", _GPU_PRESETS)
        if spec is None or spec.loader is None:
            raise RuntimeError(f"无法加载显卡型号表: {_GPU_PRESETS}")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _gpu_module = mod
    return _gpu_module


def supported_gpu_types() -> list[str]:
    """上传表单可选显卡型号（与 gpu_memory_presets.GPU_MEMORY_GIB 一致）。"""
    return sorted(_load_gpu_presets().GPU_MEMORY_GIB.keys())

# 上传体积上限（防止超大文件打满内存/磁盘）
MAX_CSV_BYTES = 5 * 1024 * 1024
MAX_XLSX_BYTES = 10 * 1024 * 1024   # xlsx 为压缩包，放宽到 10MB
MAX_TXT_BYTES = 64 * 1024


class UploadError(Exception):
    """上传内容非法。上层转成 400 反馈给用户（可修正后重试）。"""


def _decode(raw: bytes, label: str, limit: int) -> str:
    if not raw or not raw.strip():
        raise UploadError(f"{label} 为空")
    if len(raw) > limit:
        raise UploadError(f"{label} 超过大小上限（{limit // 1024} KB）")
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    raise UploadError(f"{label} 编码无法识别（请用 UTF-8 保存）")


def _is_blank_cell(c) -> bool:
    return c is None or (isinstance(c, str) and c.strip() == "")


def _norm_cell(v):
    """Excel 单元格值归一：None→''；int/float/str 原样；日期等其它类型转字符串。"""
    if v is None:
        return ""
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def _looks_like_xlsx(raw: bytes) -> bool:
    """xlsx = zip 包，固定以本地文件头魔数 PK\\x03\\x04 开头。"""
    return raw[:4] == b"PK\x03\x04"


def _looks_like_xls(raw: bytes) -> bool:
    """旧版 .xls = OLE2 复合文档，魔数 D0 CF 11 E0。"""
    return raw[:4] == b"\xd0\xcf\x11\xe0"


def _read_xlsx_table(raw: bytes) -> tuple[list[str], list[dict]]:
    """
    读取 .xlsx 首个（且必须唯一）sheet → (表头列表, 行dict列表)。
    行 dict 的 key 为去空白后的表头字符串，与 CSV 分支保持一致。
    仅支持单 sheet；多 sheet / 无法解析 / 无表头一律报错（格式不兼容）。
    """
    if len(raw) > MAX_XLSX_BYTES:
        raise UploadError(f"Excel 文件超过大小上限（{MAX_XLSX_BYTES // 1024 // 1024} MB）")
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # 运行时依赖已含 openpyxl，这里只是兜底
        raise UploadError("服务器缺少 openpyxl，无法解析 Excel，请改用 CSV") from e

    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001  非法/损坏 xlsx
        raise UploadError(f"Excel 文件无法解析（需标准 .xlsx）：{e}") from e

    try:
        sheets = wb.sheetnames
        if len(sheets) != 1:
            raise UploadError(
                f"Excel 仅支持单个 sheet，当前有 {len(sheets)} 个：{sheets}。"
                "请只保留结果所在的一个工作表后重新上传。")
        ws = wb[sheets[0]]

        header_cols: list[tuple[int, str]] = []
        fieldnames: list[str] = []
        rows: list[dict] = []
        for cells in ws.iter_rows(values_only=True):
            if not header_cols:
                # 跳过顶部整行空白；首个有值的行作为表头
                if cells is None or all(_is_blank_cell(c) for c in cells):
                    continue
                header_cols = [(i, str(c).strip())
                               for i, c in enumerate(cells) if not _is_blank_cell(c)]
                fieldnames = [name for _, name in header_cols]
                continue
            if cells is None or all(_is_blank_cell(c) for c in cells):
                continue  # 跳过空行
            record: dict = {}
            for i, name in header_cols:
                record[name] = _norm_cell(cells[i] if i < len(cells) else None)
            rows.append(record)
    finally:
        wb.close()

    if not fieldnames:
        raise UploadError("Excel 无表头（首个 sheet 为空）")
    return fieldnames, rows


def read_upload_table(raw: bytes) -> tuple[list[str], list[dict]]:
    """
    统一把上传文件读成 (表头列表, 行dict列表)，按内容自动识别 CSV / xlsx：
      · xlsx（PK\\x03\\x04 开头）→ openpyxl 解析（仅单 sheet）
      · 旧版 .xls               → 明确报错，提示另存为 xlsx / CSV
      · 其它                     → 按 CSV 文本解析
    两分支的行 dict key 统一去首尾空白，保证下游列名映射一致。
    """
    if _looks_like_xlsx(raw):
        return _read_xlsx_table(raw)
    if _looks_like_xls(raw):
        raise UploadError("暂不支持旧版 .xls，请在 Excel 中另存为 .xlsx 或导出 CSV 后重新上传")

    text = _decode(raw, "CSV 文件", MAX_CSV_BYTES)
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise UploadError("CSV 无表头")
    fieldnames = [h.strip() if isinstance(h, str) else h for h in reader.fieldnames]
    rows = [{(k.strip() if isinstance(k, str) else k): v for k, v in r.items()}
            for r in reader]
    return fieldnames, rows


def _build_metrics(fieldnames: list[str], rows: list[dict]) -> list[dict]:
    """
    表头 + 行数据 → metric 记录列表（CSV / Excel 共用）。
    表头会 remap 到 to_csv.py 规范列名；列名归一与数值转换与 scanner 一致。
    """
    if not fieldnames:
        raise UploadError("表格无表头")

    header_map = build_header_map(list(fieldnames))
    missing = check_required_dimensions(header_map)
    if missing:
        raw_headers = [str(h).strip() for h in fieldnames
                       if h is not None and str(h).strip()]
        raise UploadError(
            f"表格缺少必需列 {missing}；当前表头: {raw_headers}。"
            "请确认含 Input Length/Input_Length 与 Concurrency，或改用 to_csv.py 产出。"
        )

    remapped = format_mapping_summary(header_map)
    if remapped:
        log.info("表格列名已自动映射", extra={"fields": {"mapping": remapped}})

    metrics: list[dict] = []
    for lineno, row in enumerate(rows, start=2):
        record: dict = {}
        for col, raw in row.items():
            if col is None:
                continue
            col = str(col).strip()
            if not col:
                continue
            canon = header_map.get(col)
            if canon is None:
                continue
            key = canon.lower() if canon in schema.METRIC_DIMENSION_KEYS else canon
            record[key] = _to_number(raw.strip() if isinstance(raw, str) else raw)
        for dim in ("input_length", "concurrency"):
            if record.get(dim) is None:
                raise UploadError(f"表格第 {lineno} 行的 {dim} 为空或非数值")
        metrics.append(record)

    if not metrics:
        raise UploadError("表格无数据行")
    return metrics


def parse_table(raw: bytes) -> list[dict]:
    """上传文件字节 → metric 记录列表（自动识别 CSV / xlsx）。"""
    fieldnames, rows = read_upload_table(raw)
    return _build_metrics(fieldnames, rows)


def detect_bench_framework(csv_bytes: bytes) -> dict:
    """
    仅凭表格的 spec decoding 列是否有值，粗判 bench_framework（供前端预填）。
    支持 CSV 与 .xlsx（自动识别）。

    规则（与用户约定一致）：
      · 只有 vLLM_* spec 列有值   → 建议 vllm
      · 只有 SGLang_* spec 列有值 → 建议 sglang
      · 两者都有 / 都无           → None（交由用户手填）
    最终以用户提交的表单为准；此处不做强校验，返回诊断信息即可。
    """
    fieldnames, rows = read_upload_table(csv_bytes)
    if not fieldnames:
        raise UploadError("表格无表头")

    header_map = build_header_map(list(fieldnames))
    # 规范列名 → 表头（去空白后，与 rows 的 key 一致）
    canon_to_raw: dict[str, str] = {}
    for raw, canon in header_map.items():
        if canon and canon not in canon_to_raw:
            canon_to_raw[canon] = raw

    def any_value(canon_cols) -> bool:
        raws = [canon_to_raw[c] for c in canon_cols if c in canon_to_raw]
        if not raws:
            return False
        return any(_cell_has_value(row.get(rc)) for row in rows for rc in raws)

    sgl_has = any_value(SPEC_COLUMNS["sglang"])
    vllm_has = any_value(SPEC_COLUMNS["vllm"])
    if vllm_has and not sgl_has:
        suggestion = "vllm"
    elif sgl_has and not vllm_has:
        suggestion = "sglang"
    else:
        suggestion = None

    return {
        "bench_framework": suggestion,
        "sglang_spec_present": sgl_has,
        "vllm_spec_present": vllm_has,
    }


def parse_launch_txt(text: str) -> str:
    """
    从启动命令文本提取命令原文。
    允许 # 注释行与空行；其余非空行拼接为一条命令
    （便于直接粘贴带反斜杠换行的多行命令）。
    """
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        lines.append(s.rstrip("\\").strip())  # 去掉续行反斜杠
    cmd = " ".join(lines).strip()
    if not cmd:
        raise UploadError("启动命令无有效内容（除注释与空行外为空）")
    return cmd


def validate_meta(form: dict) -> dict:
    """校验表单元信息，返回清洗后的 dict。"""
    meta = {}
    for key in REQUIRED_META:
        val = (form.get(key) or "").strip()
        if not val:
            raise UploadError(f"缺少必填字段: {key}")
        meta[key] = val

    # 模型版本可选：无版本时存空串，库表仍保留该列以便与落盘流对齐
    meta["model_version"] = (form.get("model_version") or "").strip()

    allowed = launch_params.supported_frameworks()
    if meta["framework"] not in allowed:
        raise UploadError(f"framework 必须是 {allowed} 之一，收到: {meta['framework']}")

    gpus = supported_gpu_types()
    if meta["gpu_type"] not in gpus:
        raise UploadError(f"gpu_type 必须是 {gpus} 之一，收到: {meta['gpu_type']}")

    # ── bench 维度：必填，且与 server framework 相互独立 ──
    # bench_framework 可由 CSV 的 spec 列自动预填，但仍需随表单显式提交。
    bench_fw = (form.get("bench_framework") or "").strip()
    if not bench_fw:
        raise UploadError("缺少必填字段: bench_framework（压测工具框架，可从 CSV 自动识别或手选）")
    if bench_fw not in SUPPORTED_BENCH_FRAMEWORKS:
        raise UploadError(
            f"bench_framework 必须是 {SUPPORTED_BENCH_FRAMEWORKS} 之一，收到: {bench_fw}")
    meta["bench_framework"] = bench_fw

    # bench_flush_cache 无法从 CSV 推断，必须由用户手动勾选
    flush = _parse_bool_form(form.get("bench_flush_cache"))
    if flush is None:
        raise UploadError("请手动勾选压测前是否 flush cache（bench_flush_cache 必填，无法从 CSV 推断）")
    meta["bench_flush_cache"] = flush

    # prefix_rate 必填：本次压测共享前缀占输入长度的比例（0~1），与 input_length/concurrency
    # 同为可比对轴，无法从 CSV 推断，需用户填写。
    meta["prefix_rate"] = _parse_prefix_rate(form.get("prefix_rate"))
    return meta


def _parse_prefix_rate(raw) -> float:
    """校验并解析共享前缀占比：必填、可转 float、取值域 [0, 1)。"""
    s = ("" if raw is None else str(raw)).strip()
    if not s:
        raise UploadError("缺少必填字段: prefix_rate（本次压测共享前缀占输入长度的比例，0~1）")
    try:
        val = float(s)
    except ValueError:
        raise UploadError(f"prefix_rate 必须是数字（0~1），收到: {s}") from None
    if not (0.0 <= val < 1.0):
        raise UploadError(f"prefix_rate 必须在 [0, 1) 之间，收到: {val}")
    return val


def resolve_launch_text(launch_text: str | None) -> str:
    """从直接输入的文本解析启动命令。"""
    if not launch_text or not launch_text.strip():
        raise UploadError("请填写启动命令")
    raw = launch_text.strip()
    if len(raw.encode("utf-8")) > MAX_TXT_BYTES:
        raise UploadError(f"启动命令超过大小上限（{MAX_TXT_BYTES // 1024} KB）")
    return parse_launch_txt(raw)


def _build_pd(framework: str, prefill_text: str, decode_text: str,
              router_text: str | None) -> tuple[str, dict]:
    """
    解析 PD 分离的 prefill / decode / router 三条命令，返回 (combined_launch_cmd, pd_meta)。
    prefill / decode 必填且角色需匹配；router 可选。
    """
    prefill_cmd = resolve_launch_text(prefill_text)
    decode_cmd = resolve_launch_text(decode_text)
    router_cmd = ""
    if router_text and router_text.strip():
        if len(router_text.encode("utf-8")) > MAX_TXT_BYTES:
            raise UploadError(f"router 命令超过大小上限（{MAX_TXT_BYTES // 1024} KB）")
        router_cmd = parse_launch_txt(router_text)

    pf = launch_params.extract_role(framework, prefill_cmd)
    dc = launch_params.extract_role(framework, decode_cmd)

    if pf["role"] not in ("prefill", "both"):
        raise UploadError(
            "prefill 命令里未检测到 prefill 角色："
            "sglang 需含 --disaggregation-mode prefill，vllm 需 kv_role=kv_producer/kv_both。"
        )
    if dc["role"] not in ("decode", "both"):
        raise UploadError(
            "decode 命令里未检测到 decode 角色："
            "sglang 需含 --disaggregation-mode decode，vllm 需 kv_role=kv_consumer/kv_both。"
        )

    router = launch_params.parse_router(router_cmd)
    transfer_backend = (pf["disagg"].get("transfer_backend")
                        or dc["disagg"].get("transfer_backend"))

    pf_gpus, dc_gpus, total_gpus = launch_params.pd_gpu_counts(
        framework, pf["params"], dc["params"])

    combined = f"# PREFILL\n{prefill_cmd}\n\n# DECODE\n{decode_cmd}"
    if router_cmd:
        combined += f"\n\n# ROUTER\n{router_cmd}"

    pd_meta = {
        "transfer_backend": transfer_backend,
        "gpu_count": total_gpus,
        "prefill_gpu_count": pf_gpus,
        "decode_gpu_count": dc_gpus,
        "prefill": {
            "params": pf["params"], "launch_cmd": prefill_cmd,
            "disagg": pf["disagg"], "unrecognized": pf["unrecognized"],
            "gpu_count": pf_gpus,
        },
        "decode": {
            "params": dc["params"], "launch_cmd": decode_cmd,
            "disagg": dc["disagg"], "unrecognized": dc["unrecognized"],
            "gpu_count": dc_gpus,
        },
        "router": {**router, "launch_cmd": router_cmd},
    }
    return combined, pd_meta


def ingest(
    db,
    meta_form: dict,
    csv_bytes: bytes,
    launch_text: str | None = None,
    *,
    benchmark_root: str,
    dir_pattern: str,
    deployment_mode: str = "colocated",
    prefill_text: str | None = None,
    decode_text: str | None = None,
    router_text: str | None = None,
) -> dict:
    """
    完整上传流程：校验 → 解析 → 落盘 → 扫描入库。
    落盘后立即执行一轮 Scanner，与独立 Scanner 进程逻辑一致。

    deployment_mode:
      'colocated' —— 单机/分布式：读 launch_text 一条命令（原有行为）；
      'pd_disagg' —— PD 分离：读 prefill_text / decode_text（必填）+ router_text（可选）。
    """
    if deployment_mode not in ("colocated", "pd_disagg"):
        raise UploadError(f"deployment_mode 非法: {deployment_mode}")

    meta = validate_meta(meta_form)
    metrics = parse_table(csv_bytes)  # 自动识别 CSV / .xlsx

    extra: dict = {"ingest_source": "manual_upload"}

    if deployment_mode == "pd_disagg":
        launch_cmd, pd_meta = _build_pd(
            meta["framework"], prefill_text, decode_text, router_text)
        params: dict = {}
        pd_block: dict | None = pd_meta
        extra["gpu_count"] = pd_meta["gpu_count"]
        extra["prefill_gpu_count"] = pd_meta["prefill_gpu_count"]
        extra["decode_gpu_count"] = pd_meta["decode_gpu_count"]
    else:
        launch_cmd = resolve_launch_text(launch_text)
        # 安全网：单机模式却贴了 PD 命令 → 提示改用 PD 分离（前端应已自动跳转）
        if launch_params.looks_like_pd(launch_cmd):
            raise UploadError(
                "检测到 disaggregation / kv-transfer-config 相关参数，"
                "请切换到「PD 分离」分别填写 prefill 与 decode 命令。"
            )
        params, base_extra = launch_params.extract(meta["framework"], launch_cmd)
        extra.update(base_extra)
        pd_block = None

    now = datetime.now(timezone.utc)
    try:
        dir_name, dir_path = persist.persist_upload(
            benchmark_root, db, now, meta, metrics, launch_cmd, params, extra,
            deployment_mode=deployment_mode, pd=pd_block,
        )
    except persist.PersistError as e:
        raise UploadError(str(e)) from e

    from autores.scanner.main import scan_once

    pending, ok, fail = scan_once(db, benchmark_root, dir_pattern)
    runs = db.fetch_runs("run_id = ?", [dir_name])
    if not runs:
        raise UploadError(
            f"已落盘至 {dir_path}，但扫描入库失败（pending={pending}, ok={ok}, fail={fail}）"
        )
    doc = runs[0]
    log.info("上传落盘并扫描入库", extra={"fields": {
        "dir": dir_path, "run_id": doc["_id"], "deployment_mode": deployment_mode}})

    summary = {
        "run_id": doc["_id"],
        "source_dir": dir_name,
        "disk_path": dir_path,
        "num_metrics": len(doc["metrics"]),
        "deployment_mode": deployment_mode,
        "gpu_count": doc.get("gpu_count"),
        "launch_cmd": doc["launch_cmd"],
    }
    if deployment_mode == "pd_disagg":
        summary["pd"] = {
            "transfer_backend": pd_block["transfer_backend"],
            "gpu_count": pd_block["gpu_count"],
            "prefill_gpu_count": pd_block["prefill_gpu_count"],
            "decode_gpu_count": pd_block["decode_gpu_count"],
            "prefill": {"params": pd_block["prefill"]["params"],
                        "unrecognized": pd_block["prefill"]["unrecognized"]},
            "decode": {"params": pd_block["decode"]["params"],
                       "unrecognized": pd_block["decode"]["unrecognized"]},
            "router": {k: v for k, v in pd_block["router"].items() if k != "launch_cmd"},
        }
    else:
        summary["params"] = doc.get("params", {})
        summary["unrecognized"] = doc.get("extra", {}).get("unrecognized", [])
    return summary
